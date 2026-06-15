#!/usr/bin/env python3
"""
Convert toy problem Excel files to gapfill data format.

This module processes Excel files containing flux balance analysis (FBA) toy problem
data and converts them to a simplified data format suitable for metabolic gap-filling
analysis. The output consists of reaction names with flux values and error margins.

The expected Excel format includes columns for:
- Reaction names
- Direction indicators (>, <, =)
- Flux values (numeric or 'max' for biomass)
- Compartment information (from/to)

Examples
--------
Command line usage:
    $ python toy2dat.py input.xlsx > output.dat
    $ python toy2dat.py -f 7 input.xlsx > output.dat

Notes
-----
Standard compartment name combinations for transport reactions are defined
in the global `std_names` list: ["EP", "PC", "CE"].

Output format is: reaction_name flux error
Special values:
- flux = -1.0, error = 0.0: biomass equation (objective function)
- flux = 0.0, error = 9999.0: unknown flux value
"""

import sys
import os
import fileinput
import openpyxl

std_names = ["EP", "PC", "CE"]

def die (message):
    """
    Print error message to stderr and exit with status 1.

    Parameters
    ----------
    message : str
        Error message to display before terminating the program.

    Returns
    -------
    None
        This function does not return; it terminates the program.

    Examples
    --------
    >>> die("Invalid file format")
    Invalid file format
    (program exits)
    """
    print (message, file=sys.stderr)
    exit(1)

def usage (str = ""):
    """
    Display usage information and exit.

    Prints the command-line usage instructions including available options
    and arguments, then terminates the program with exit status 1.

    Parameters
    ----------
    str : str, optional
        Additional message to display before usage information (default: "").

    Returns
    -------
    None
        This function does not return; it terminates the program.

    Notes
    -----
    The function displays:
    - Optional custom message
    - Command syntax
    - Description of the conversion process
    - Required arguments (input Excel file)
    - Optional switches (-f for flux column specification)
    """
    print (str)
    print (sys.argv[0], end=' ')
    print ('''[-f #] in.xlsx 
        Convert a toy prob xlsx into gapfill data
        Writes to stdout
    Args
        file to process
    Switches
        -f: column for flux (1-based) [6]
''')
    exit(1)

def main ():
    """
    Main entry point for toy2dat conversion.

    Parses command-line arguments, reads the input Excel file, and converts
    flux balance analysis data to gapfill format. Processes each reaction row
    and outputs formatted data to stdout.

    Parameters
    ----------
    None

    Returns
    -------
    None

    Raises
    ------
    SystemExit
        If required arguments are missing or invalid.
    Exception
        If Excel file cannot be loaded or processed.

    Notes
    -----
    Command-line arguments:
    - Input file: Excel (.xlsx) file containing toy problem data
    - -f #: Specify flux column number (1-based indexing, default: 6)

    Expected Excel structure (starting at row 2):
    - Column 0: Reaction name
    - Column flux_col-1: Direction (>, <, =)
    - Column flux_col: Flux value (number, 'max', or empty)
    - Column flux_col+1: To compartment
    - Column flux_col+2: From compartment

    Output format (written to stdout):
    - For biomass (flux='max'): reaction_compartment -1.0 0.0
    - For unknown flux: reaction_compartment 0.0 9999.0
    - For known flux: reaction_compartment flux error

    Transport reactions are identified when from_compart != to_compart,
    and compartment names are validated against std_names.

    Reverse reactions (indicated by '<' direction or negative flux) are
    suffixed with '_rev'.

    Examples
    --------
    Process default columns:
        $ python toy2dat.py experiment.xlsx > output.dat

    Specify custom flux column (column 7):
        $ python toy2dat.py -f 7 experiment.xlsx > output.dat
    """
    global std_names
    flux_col = 5
    dir_col = 4
    filename = ""
    while len(sys.argv) > 1:
        arg = sys.argv.pop(1)
        if (arg[0] == '-'):
            if arg == '-f' and len(sys.argv) > 1:
                flux_col = int(sys.argv.pop(1)) - 1
            else:
                usage("Unrecognised arg: " + arg)
        else:
            # Save for file processing
            if filename == "":
                filename = arg
            else:
                usage ("Too many filenames")
                
    if filename == "":
        usage ("Filename required")

    to_compart_col = flux_col + 1
    from_compart_col = to_compart_col + 1
    dir_col = flux_col - 1
    
    wb = openpyxl.load_workbook(filename = filename)
    ws = wb.active
    
    for row in range(2,ws.max_row+1):
        # Only write out real reactions
        if ws[row] == None or \
                len (ws[row]) == 0 or \
                ws[row][0].value == None:
            continue
        name = ws[row][0].value
        flux_fld = ws[row][flux_col].value
        to_compart = ws[row][to_compart_col].value
        from_compart = ws[row][from_compart_col].value
        dir_fld = ws[row][dir_col].value
        #if not flux_fld is None:
        #    print ("Flux >" + flux_fld + "<")

        compart = from_compart
        if from_compart != to_compart:
            # Its a transport
            compart = from_compart + to_compart
            if compart not in std_names:
                compart = to_compart + from_compart
            if compart not in std_names:
                die ("Illegal compart comb: " + from_compart + to_compart)

        if flux_fld == "max":
            # This is the biomass equation
            if dir_fld == '<':
                name = name + '_rev'
            print (name + "_" + compart, -1.0, 0.0)
        elif flux_fld is None or len(flux_fld) == 0:
            # Unknown flux
            # Output a 0 flux with special 9999 error, indicating unknown
            if dir_fld == '>' or dir_fld == '=':
                print (name + "_" + compart, 0.0, 9999.0)
            # Also/instead put out reverse flow if req'd
            if dir_fld == '<' or dir_fld == '=':
                print (name + '_rev' + "_" + compart, 0.0, 9999.0)
        else:
            fld_tok = flux_fld.split ('±')
            flux = float (fld_tok[0])
            error = float (fld_tok[1])
            if flux < 0:
                # Reversed flow
                flux = -flux
                name = name + '_rev'
            print (name + "_" + compart, flux, error)
                
    
if __name__ == '__main__':
    main()
